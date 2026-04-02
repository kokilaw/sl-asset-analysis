#!/usr/bin/env python3
"""
Monthly Unit Trust Report

Run:
    python3 portfolio_monthly.py

Outputs:
    generated/monthly_portfolio_report.html
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import os
import re
from difflib import SequenceMatcher

import requests
from openpyxl import load_workbook

from portfolio import CAL_UT_API


def get_google_sheet_xlsx_url() -> str:
    """Resolve Google Sheet XLSX URL from environment variables."""
    direct_url = (os.getenv("GOOGLE_SHEET_XLSX_URL") or "").strip()
    if direct_url:
        return direct_url

    sheet_id = (os.getenv("GOOGLE_SHEET_ID") or "").strip()
    if sheet_id:
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    raise RuntimeError(
        "Missing Google Sheet config. Set GOOGLE_SHEET_XLSX_URL or GOOGLE_SHEET_ID."
    )


def fmt_lkr(value, decimals=2):
    if value is None:
        return "N/A"
    return f"LKR {value:,.{decimals}f}"


def fmt_pct(value, decimals=2):
    if value is None:
        return "N/A"
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.{decimals}f}%"


def pnl_class(value):
    if value is None:
        return "neutral"
    return "positive" if value >= 0 else "negative"


def pnl_arrow(value):
    if value is None:
        return ""
    return "▲" if value >= 0 else "▼"


def _to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "":
        return None
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except ValueError:
        return None


def _to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def fetch_ut_navs() -> dict:
    """Get latest NAV per fund from CAL API."""
    response = requests.get(CAL_UT_API, timeout=20)
    response.raise_for_status()
    raw = response.json()

    navs = {}
    for row in raw:
        name = (row.get("fundname") or "").strip()
        sell = _to_number(row.get("sell"))
        nav_date = row.get("date")
        if name and sell is not None:
            navs[name] = {"nav": sell, "date": nav_date}
    return navs


def _best_fund_match(sheet_name: str, api_funds: list[str]) -> str | None:
    if not api_funds:
        return None

    ns = _norm(sheet_name)

    exact_candidates = []
    for fund in api_funds:
        nf = _norm(fund)
        if ns == nf or ns.startswith(nf) or nf.startswith(ns):
            exact_candidates.append(fund)
    if len(exact_candidates) == 1:
        return exact_candidates[0]
    if len(exact_candidates) > 1:
        exact_candidates.sort(key=lambda f: len(_norm(f)), reverse=True)
        return exact_candidates[0]

    scored = sorted(
        ((SequenceMatcher(None, ns, _norm(f)).ratio(), f) for f in api_funds),
        reverse=True,
    )
    return scored[0][1] if scored and scored[0][0] >= 0.55 else None


def fetch_ut_transactions_from_sheet() -> list[dict]:
    """Read all worksheet tabs from the Google Sheet and parse transactions."""
    sheet_url = get_google_sheet_xlsx_url()
    response = requests.get(sheet_url, timeout=30)
    response.raise_for_status()

    wb = load_workbook(filename=BytesIO(response.content), data_only=True, read_only=True)

    all_funds = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        headers = None
        txns = []

        for row in rows:
            if row is None or not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            if headers is None:
                normalized = []
                for cell in row:
                    h = str(cell).strip().lower() if cell is not None else ""
                    h = h.replace(" ", "")
                    normalized.append(h)

                required = {"date", "type", "units", "unitprice", "total"}
                if required.issubset(set(normalized)):
                    headers = {name: idx for idx, name in enumerate(normalized)}
                continue

            try:
                dt = _to_date(row[headers["date"]])
                tx_type = str(row[headers["type"]]).strip() if row[headers["type"]] is not None else ""
                units = _to_number(row[headers["units"]])
                unit_price = _to_number(row[headers["unitprice"]])
                total = _to_number(row[headers["total"]])
            except Exception:
                continue

            # Normalize cashflow sign:
            # - Investments should be positive contributions
            # - Redemptions should be negative contributions
            # If the sheet's Total sign is inconsistent, infer from Units/Type.
            if total is not None:
                t = tx_type.lower()
                if units is not None:
                    if units < 0 and total > 0:
                        total = -total
                    elif units > 0 and total < 0:
                        total = abs(total)
                elif "redeem" in t and total > 0:
                    total = -total
                elif "invest" in t and total < 0:
                    total = abs(total)

            if dt is None and units is None and total is None:
                continue

            txns.append(
                {
                    "date": dt,
                    "type": tx_type,
                    "units": units or 0.0,
                    "unit_price": unit_price,
                    "total": total or 0.0,
                }
            )

        if txns:
            all_funds.append({"sheet_name": ws.title, "transactions": txns})

    return all_funds


def build_ut_holdings() -> tuple[list[dict], dict]:
    nav_map = fetch_ut_navs()
    nav_funds = list(nav_map.keys())
    funds = fetch_ut_transactions_from_sheet()

    today = date.today()
    month_start = date(today.year, today.month, 1)
    holdings = []
    recent_activity = []

    for fund in funds:
        sheet_name = fund["sheet_name"]
        txns = sorted(
            fund["transactions"],
            key=lambda t: t.get("date") or date.min,
        )
        matched_name = _best_fund_match(sheet_name, nav_funds)

        units_held = sum(t["units"] for t in txns)
        net_contrib = sum(t["total"] for t in txns)
        last_txn = max((t["date"] for t in txns if t.get("date")), default=None)

        # Remaining cost basis using moving-average method.
        running_units = 0.0
        running_cost = 0.0
        realized_pnl = 0.0
        opening_units = 0.0
        opening_cost = 0.0

        for t in txns:
            units = t["units"]
            cash = t["total"]
            unit_price = t.get("unit_price") or 0.0
            tx_date = t.get("date")

            if units > 0:
                cost_in = abs(cash) if cash != 0 else units * unit_price
                running_units += units
                running_cost += cost_in
            elif units < 0:
                sold_units = abs(units)
                avg_cost = (running_cost / running_units) if running_units > 0 else 0.0
                cost_out = avg_cost * sold_units
                proceeds = abs(cash) if cash != 0 else sold_units * unit_price

                running_units += units
                running_cost -= cost_out
                if running_cost < 0:
                    running_cost = 0.0

                realized_pnl += proceeds - cost_out

            if tx_date and tx_date < month_start:
                opening_units = running_units
                opening_cost = running_cost

            if tx_date and tx_date.year == today.year and tx_date.month == today.month:
                recent_activity.append(
                    {
                        "date": tx_date,
                        "fund_name": matched_name or sheet_name,
                        "type": t.get("type") or "",
                        "units": units,
                        "total": cash,
                    }
                )

        month_flow = sum(
            t["total"]
            for t in txns
            if t.get("date") and t["date"].year == today.year and t["date"].month == today.month
        )
        month_txn_count = sum(
            1
            for t in txns
            if t.get("date") and t["date"].year == today.year and t["date"].month == today.month
        )

        nav = None
        nav_date = None
        if matched_name and matched_name in nav_map:
            nav = nav_map[matched_name]["nav"]
            nav_date = nav_map[matched_name]["date"]

        nav_date_obj = _to_date(nav_date)
        nav_age_days = (today - nav_date_obj).days if nav_date_obj else None

        current_value = units_held * nav if nav is not None else None
        opening_value = opening_units * nav if nav is not None else None
        month_pnl_est = (
            (current_value - opening_value - month_flow)
            if (current_value is not None and opening_value is not None)
            else None
        )
        month_return_est = (
            (month_pnl_est / opening_value * 100)
            if (month_pnl_est is not None and opening_value and opening_value > 0)
            else None
        )
        unrealized_pnl = (current_value - running_cost) if current_value is not None else None
        unrealized_pct = (
            unrealized_pnl / running_cost * 100
            if (unrealized_pnl is not None and running_cost > 0)
            else None
        )
        breakeven_nav = (running_cost / units_held) if units_held else None

        holdings.append(
            {
                "sheet_name": sheet_name,
                "fund_name": matched_name or sheet_name,
                "matched_name": matched_name,
                "units": round(units_held, 2),
                "net_contribution": round(net_contrib, 2),
                "remaining_cost": round(running_cost, 2),
                "realized_pnl": round(realized_pnl, 2),
                "breakeven_nav": round(breakeven_nav, 4) if breakeven_nav is not None else None,
                "current_nav": nav,
                "nav_date": nav_date,
                "nav_age_days": nav_age_days,
                "current_value": round(current_value, 2) if current_value is not None else None,
                "pnl_lkr": round(unrealized_pnl, 2) if unrealized_pnl is not None else None,
                "pnl_pct": round(unrealized_pct, 2) if unrealized_pct is not None else None,
                "last_txn": last_txn,
                "month_flow": round(month_flow, 2),
                "month_pnl_est": round(month_pnl_est, 2) if month_pnl_est is not None else None,
                "month_return_est": round(month_return_est, 2) if month_return_est is not None else None,
                "opening_value": round(opening_value, 2) if opening_value is not None else None,
                "month_txn_count": month_txn_count,
                "txn_count": len(txns),
            }
        )

    valid = [h for h in holdings if h.get("current_value") is not None]
    holdings.sort(key=lambda h: h.get("current_value") or 0, reverse=True)
    total_cost_basis = sum(h["remaining_cost"] for h in holdings)
    total_contrib = sum(h["net_contribution"] for h in holdings)
    total_value = sum(h["current_value"] for h in valid)
    total_pnl = total_value - total_cost_basis if valid else None
    total_pnl_pct = (total_pnl / total_cost_basis * 100) if (total_pnl is not None and total_cost_basis) else None
    total_month_flow = sum(h["month_flow"] for h in holdings)
    total_month_pnl_est = sum((h.get("month_pnl_est") or 0) for h in valid)
    total_opening_value = sum((h.get("opening_value") or 0) for h in valid)
    total_month_return_est = (
        total_month_pnl_est / total_opening_value * 100
        if total_opening_value > 0
        else None
    )
    total_realized_pnl = sum(h["realized_pnl"] for h in holdings)
    active_funds_this_month = sum(1 for h in holdings if h.get("month_txn_count", 0) > 0)
    missing_nav_count = sum(1 for h in holdings if h.get("current_nav") is None)
    stale_nav_count = sum(1 for h in holdings if h.get("nav_age_days") is not None and h["nav_age_days"] > 7)
    unmatched_count = sum(1 for h in holdings if not h.get("matched_name"))

    for h in holdings:
        cv = h.get("current_value") or 0
        h["allocation_pct"] = round(cv / total_value * 100, 1) if total_value else 0
        if total_pnl and h.get("pnl_lkr") is not None:
            h["pnl_contribution_pct"] = round(h["pnl_lkr"] / total_pnl * 100, 1)
        else:
            h["pnl_contribution_pct"] = None

    best = max(valid, key=lambda x: x.get("pnl_pct", -10**9), default=None)
    worst = min(valid, key=lambda x: x.get("pnl_pct", 10**9), default=None)

    recent_activity.sort(key=lambda x: x["date"], reverse=True)

    summary = {
        "fund_count": len(holdings),
        "valued_funds": len(valid),
        "total_cost_basis": round(total_cost_basis, 2),
        "total_contribution": round(total_contrib, 2),
        "total_value": round(total_value, 2),
        "total_pnl": round(total_pnl, 2) if total_pnl is not None else None,
        "total_pnl_pct": round(total_pnl_pct, 2) if total_pnl_pct is not None else None,
        "total_month_flow": round(total_month_flow, 2),
        "total_month_pnl_est": round(total_month_pnl_est, 2),
        "total_month_return_est": round(total_month_return_est, 2) if total_month_return_est is not None else None,
        "total_realized_pnl": round(total_realized_pnl, 2),
        "active_funds_this_month": active_funds_this_month,
        "missing_nav_count": missing_nav_count,
        "stale_nav_count": stale_nav_count,
        "unmatched_count": unmatched_count,
        "recent_activity": recent_activity[:12],
        "best": best,
        "worst": worst,
    }

    return holdings, summary


def generate_monthly_html(ut_holdings: list, ut_summary: dict) -> str:
        report_date = datetime.now().strftime("%B %d, %Y at %H:%M")
        today = date.today().isoformat()

        ut_total_value = ut_summary["total_value"]
        ut_total_cost = ut_summary["total_cost_basis"]
        ut_total_pnl = ut_summary["total_pnl"]
        ut_total_pct = ut_summary["total_pnl_pct"]

        def card(title: str, value: str, value_class: str = "") -> str:
                return (
                        f"<div class=\"card\">"
                        f"<div class=\"card-label\">{title}</div>"
                        f"<div class=\"card-value {value_class}\">{value}</div>"
                        f"</div>"
                )

        def health_badge(label: str, value: str, cls: str = "neutral") -> str:
                return f"<span class=\"health-badge {cls}\"><strong>{label}:</strong> {value}</span>"

        cards_html = "".join(
                [
                        card("Current Value", fmt_lkr(ut_total_value), "bold"),
                        card("Cost Basis", fmt_lkr(ut_total_cost)),
                        card(
                                "Unrealized P&L",
                                f"{pnl_arrow(ut_total_pnl)} {fmt_lkr(ut_total_pnl)} ({fmt_pct(ut_total_pct)})"
                                if ut_total_pnl is not None
                                else "N/A",
                                pnl_class(ut_total_pnl),
                        ),
                        card(
                                "Realized P&L",
                                fmt_lkr(ut_summary["total_realized_pnl"]),
                                pnl_class(ut_summary["total_realized_pnl"]),
                        ),
                        card(
                                "Net Flow (This Month)",
                                fmt_lkr(ut_summary["total_month_flow"]),
                                pnl_class(ut_summary["total_month_flow"]),
                        ),
                        card(
                                "Estimated Month Return",
                                (
                                        f"{pnl_arrow(ut_summary['total_month_pnl_est'])} {fmt_lkr(ut_summary['total_month_pnl_est'])} "
                                        f"({fmt_pct(ut_summary['total_month_return_est'])})"
                                )
                                if ut_summary.get("total_month_return_est") is not None
                                else fmt_lkr(ut_summary["total_month_pnl_est"]),
                                pnl_class(ut_summary.get("total_month_pnl_est")),
                        ),
                ]
        )

        health_html = " ".join(
                [
                        health_badge("Funds (valued/total)", f"{ut_summary['valued_funds']}/{ut_summary['fund_count']}", "neutral"),
                        health_badge("Active this month", str(ut_summary["active_funds_this_month"]), "neutral"),
                        health_badge(
                                "Missing NAV",
                                str(ut_summary["missing_nav_count"]),
                                "negative" if ut_summary["missing_nav_count"] > 0 else "positive",
                        ),
                        health_badge(
                                "Stale NAV (>7d)",
                                str(ut_summary["stale_nav_count"]),
                                "negative" if ut_summary["stale_nav_count"] > 0 else "positive",
                        ),
                        health_badge(
                                "Unmatched fund names",
                                str(ut_summary["unmatched_count"]),
                                "negative" if ut_summary["unmatched_count"] > 0 else "positive",
                        ),
                ]
        )

        allocation_rows = ""
        for h in ut_holdings:
                nav_age = f"{h['nav_age_days']}d" if h.get("nav_age_days") is not None else "N/A"
                allocation_rows += f"""
                <tr>
                    <td>{h['fund_name']}</td>
                    <td class=\"num\">{h['allocation_pct']:.1f}%</td>
                    <td class=\"num\">{fmt_lkr(h['current_value']) if h.get('current_value') is not None else 'N/A'}</td>
                    <td class=\"num {pnl_class(h.get('pnl_lkr'))}\">{fmt_lkr(h['pnl_lkr']) if h.get('pnl_lkr') is not None else 'N/A'}</td>
                    <td class=\"num {pnl_class(h.get('pnl_contribution_pct'))}\">{fmt_pct(h['pnl_contribution_pct'], 1) if h.get('pnl_contribution_pct') is not None else 'N/A'}</td>
                    <td class=\"num\">{nav_age}</td>
                </tr>
                """

        ut_rows = ""
        for h in ut_holdings:
                nav = f"{h['current_nav']:.4f}" if h.get("current_nav") is not None else "N/A"
                breakeven_nav = f"{h['breakeven_nav']:.4f}" if h.get("breakeven_nav") is not None else "N/A"
                last_txn = h["last_txn"].isoformat() if h.get("last_txn") else "N/A"
                nav_freshness = f"{h['nav_age_days']}d" if h.get("nav_age_days") is not None else "N/A"

                ut_rows += f"""
                <tr>
                    <td>{h['fund_name']}</td>
                    <td class=\"num\">{h['units']:,.2f}</td>
                    <td class=\"num\">{fmt_lkr(h['remaining_cost'])}</td>
                    <td class=\"num\">{breakeven_nav}</td>
                    <td class=\"num\">{nav}</td>
                    <td class=\"num\">{fmt_lkr(h['current_value']) if h.get('current_value') is not None else 'N/A'}</td>
                    <td class=\"num {pnl_class(h['pnl_lkr'])}\">{pnl_arrow(h['pnl_lkr'])} {fmt_lkr(h['pnl_lkr']) if h.get('pnl_lkr') is not None else 'N/A'}</td>
                    <td class=\"num {pnl_class(h['pnl_pct'])}\">{fmt_pct(h['pnl_pct'])}</td>
                    <td class=\"num {pnl_class(h['month_flow'])}\">{fmt_lkr(h['month_flow'])}</td>
                    <td class=\"num {pnl_class(h.get('month_pnl_est'))}\">{fmt_lkr(h['month_pnl_est']) if h.get('month_pnl_est') is not None else 'N/A'}</td>
                    <td>{nav_freshness}</td>
                    <td>{last_txn}</td>
                </tr>
                """

        activity_rows = ""
        for a in ut_summary.get("recent_activity", []):
                activity_rows += f"""
                <tr>
                    <td>{a['date'].isoformat()}</td>
                    <td>{a['fund_name']}</td>
                    <td>{a['type'] or 'N/A'}</td>
                    <td class=\"num\">{a['units']:,.2f}</td>
                    <td class=\"num {pnl_class(a['total'])}\">{fmt_lkr(a['total'])}</td>
                </tr>
                """
        if not activity_rows:
                activity_rows = """
                <tr>
                    <td colspan=\"5\" class=\"neutral\">No transactions found for this month.</td>
                </tr>
                """

        best_html = ""
        if ut_summary.get("best"):
                b = ut_summary["best"]
                w = ut_summary["worst"]
                best_html = f"""
                <p class=\"subtitle\" style=\"margin-top:-14px;margin-bottom:24px\">
                    🏆 Best UT: <strong>{b['fund_name']}</strong> ({fmt_pct(b['pnl_pct'])})
                    &nbsp;|&nbsp;
                    📉 Worst UT: <strong>{w['fund_name']}</strong> ({fmt_pct(w['pnl_pct'])})
                </p>
                """

        return f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
    <meta charset=\"UTF-8\"/>
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\"/>
    <title>Monthly Portfolio Report — {today}</title>
    <style>
        *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
                     background: #f0f4f8; color: #1e293b; padding: 24px 32px; max-width: 1450px; margin: 0 auto; }}
        h1 {{ font-size: 1.8rem; font-weight: 700; color: #0f172a; }}
        h2 {{ font-size: 1rem; color: #334155; margin-bottom: 14px; }}
        .subtitle {{ color: #64748b; font-size: 0.9rem; margin-top: 6px; margin-bottom: 24px; }}

        .cards {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 14px; margin-bottom: 24px; }}
        .card {{ background: #fff; border-radius: 12px; padding: 16px 18px; box-shadow: 0 1px 4px rgba(0,0,0,.07); }}
        .card-label {{ font-size: 0.72rem; text-transform: uppercase; letter-spacing: .06em; color: #64748b; margin-bottom: 8px; }}
        .card-value {{ font-size: 1.15rem; font-weight: 700; }}
        .card-value.bold {{ color: #0f172a; }}

        .health-row {{ display: flex; flex-wrap: wrap; gap: 10px; }}
        .health-badge {{ display: inline-flex; align-items: center; gap: 6px; padding: 6px 10px;
                                         border-radius: 999px; font-size: 0.78rem; border: 1px solid transparent; }}
        .health-badge.positive {{ color: #15803d; background: #dcfce7; border-color: #bbf7d0; }}
        .health-badge.negative {{ color: #b91c1c; background: #fee2e2; border-color: #fecaca; }}
        .health-badge.neutral  {{ color: #475569; background: #f1f5f9; border-color: #e2e8f0; }}

        .table-wrap {{ background: #fff; border-radius: 12px; padding: 18px 20px; box-shadow: 0 1px 4px rgba(0,0,0,.07); overflow-x: auto; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 0.84rem; }}
        thead tr {{ border-bottom: 2px solid #e2e8f0; }}
        thead th {{ background: #f8fafc; text-align: left; padding: 9px 10px; font-size: 0.7rem; text-transform: uppercase; letter-spacing: .05em; color: #94a3b8; white-space: nowrap; }}
        tbody tr {{ border-bottom: 1px solid #f1f5f9; }}
        tbody tr:hover {{ background: #f8fafc; }}
        td {{ padding: 9px 10px; white-space: nowrap; }}
        td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}

        .positive {{ color: #16a34a; }}
        .negative {{ color: #dc2626; }}
        .neutral {{ color: #64748b; }}

        .footer {{ margin-top: 24px; font-size: 0.76rem; color: #94a3b8; text-align: center; }}
    </style>
</head>
<body>
    <h1>🗓️ Monthly Unit Trust Report</h1>
    <p class=\"subtitle\">Generated on {report_date} · Unit trusts from Google Sheet + CAL NAV API</p>

    <div class=\"cards\">{cards_html}</div>
    {best_html}

    <div class=\"table-wrap\" style=\"padding-bottom:14px\">
        <h2>✅ Portfolio Health Checks</h2>
        <div class=\"health-row\">{health_html}</div>
    </div>

    <div class=\"table-wrap\">
        <h2>🥧 Allocation & Contribution</h2>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Fund</th>
                    <th>Type</th>
                    <th style="text-align:right">Units</th>
                    <th style="text-align:right">Cash Flow</th>
                </tr>
            </thead>
            <tbody>{activity_rows}</tbody>
        </table>
    </div>

    <div class="table-wrap">
        <h2>📘 Quick Guide (Key Terms)</h2>
        <table>
            <thead>
                <tr>
                    <th>Metric</th>
                    <th>What it means</th>
                    <th>What to look for</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>Cost Basis</td>
                    <td>Remaining invested amount for current units (after redemptions).</td>
                    <td>Used as the baseline for unrealized return.</td>
                </tr>
                <tr>
                    <td>Current Value</td>
                    <td>Current units × latest NAV.</td>
                    <td>Should grow above Cost Basis over long periods.</td>
                </tr>
                <tr>
                    <td>Unrealized P&amp;L / Unrealized %</td>
                    <td>Profit/loss if you valued holdings today without selling.</td>
                    <td>Positive is favorable; compare across funds for relative strength.</td>
                </tr>
                <tr>
                    <td>Realized P&amp;L</td>
                    <td>Profit/loss already locked in from past redemptions.</td>
                    <td>Sustainably positive over time is generally better.</td>
                </tr>
                <tr>
                    <td>Break-even NAV</td>
                    <td>NAV needed for current holding to be at zero unrealized P&amp;L.</td>
                    <td>Current NAV above break-even is favorable.</td>
                </tr>
                <tr>
                    <td>Allocation</td>
                    <td>Fund's share of total portfolio value.</td>
                    <td>Keep diversification; avoid very high concentration in one fund.</td>
                </tr>
                <tr>
                    <td>Contribution to Total P&amp;L</td>
                    <td>How much each fund drives total unrealized portfolio gain/loss.</td>
                    <td>Large positive contributors are leading returns; large negative need review.</td>
                </tr>
                <tr>
                    <td>NAV Age / Missing NAV / Stale NAV</td>
                    <td>How fresh valuation data is, and whether any values are unavailable.</td>
                    <td>Prefer low NAV age, zero missing NAV, and zero stale NAV.</td>
                </tr>
                <tr>
                    <td>This Month Flow &amp; Est. Month P&amp;L</td>
                    <td>Net cash added/withdrawn this month and estimated monthly investment result.</td>
                    <td>Read together: flow explains capital movement, P&amp;L explains performance.</td>
                </tr>
            </tbody>
        </table>
    </div>
            <thead>
                <tr>
                    <th>Fund</th>
                    <th style=\"text-align:right\">Allocation</th>
                    <th style=\"text-align:right\">Current Value</th>
                    <th style=\"text-align:right\">Unrealized P&L</th>
                    <th style=\"text-align:right\">Contribution to Total P&L</th>
                    <th style=\"text-align:right\">NAV Age</th>
                </tr>
            </thead>
            <tbody>{allocation_rows}</tbody>
        </table>
    </div>

    <div class=\"table-wrap\">
        <h2>🏦 Detailed Holdings</h2>
        <table>
            <thead>
                <tr>
                    <th>Fund</th>
                    <th style=\"text-align:right\">Units</th>
                    <th style=\"text-align:right\">Cost Basis</th>
                    <th style=\"text-align:right\">Break-even NAV</th>
                    <th style=\"text-align:right\">Current NAV</th>
                    <th style=\"text-align:right\">Current Value</th>
                    <th style=\"text-align:right\">Unrealized P&L</th>
                    <th style=\"text-align:right\">Unrealized %</th>
                    <th style=\"text-align:right\">This Month Flow</th>
                    <th style=\"text-align:right\">Est. Month P&L</th>
                    <th>NAV Age</th>
                    <th>Last Transaction</th>
                </tr>
            </thead>
            <tbody>{ut_rows}</tbody>
        </table>
    </div>

    <div class=\"table-wrap\">
        <h2>🧾 Recent Activity (Current Month)</h2>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Fund</th>
                    <th>Type</th>
                    <th style=\"text-align:right\">Units</th>
                    <th style=\"text-align:right\">Cash Flow</th>
                </tr>
            </thead>
            <tbody>{activity_rows}</tbody>
        </table>
    </div>

    <p class=\"footer\">For personal tracking only — not financial advice.</p>
</body>
</html>
"""


def main():
    print("🗓️ Monthly Unit Trust Reporter")
    print("─" * 42)

    print("Fetching unit trust transactions from Google Sheet...")
    ut_holdings, ut_summary = build_ut_holdings()

    html = generate_monthly_html(ut_holdings, ut_summary)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(base_dir, "generated")
    os.makedirs(out_dir, exist_ok=True)

    out_file = os.path.join(out_dir, "monthly_portfolio_report.html")
    with open(out_file, "w", encoding="utf-8") as f:
        f.write(html)

    print("\n✅ Monthly report generated")
    print("   File: generated/monthly_portfolio_report.html")
    print(f"   Open: file://{out_file}")


if __name__ == "__main__":
    main()
